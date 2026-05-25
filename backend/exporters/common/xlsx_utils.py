from __future__ import annotations

from typing import Iterable, Tuple

import pandas as pd

try:
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    HAS_OPENPYXL = True
except ImportError:  # pragma: no cover
    HAS_OPENPYXL = False
    Alignment = None  # type: ignore[assignment]
    Border = None  # type: ignore[assignment]
    Font = None  # type: ignore[assignment]
    PatternFill = None  # type: ignore[assignment]
    Side = None  # type: ignore[assignment]


Section = Tuple[str, pd.DataFrame]


def write_sectioned_sheet(writer, sheet_name: str, sections: Iterable[Section]) -> None:
    startrow = 0
    wrote_any = False
    for title, frame in sections:
        clean_frame = frame if isinstance(frame, pd.DataFrame) and not frame.empty else pd.DataFrame([["无数据"]], columns=["说明"])
        pd.DataFrame([[title]], columns=["章节"]).to_excel(
            writer,
            sheet_name=sheet_name,
            index=False,
            header=False,
            startrow=startrow,
        )
        clean_frame.to_excel(writer, sheet_name=sheet_name, index=False, startrow=startrow + 1)
        startrow += len(clean_frame.index) + 4
        wrote_any = True
    if not wrote_any:
        pd.DataFrame([["无数据"]], columns=["说明"]).to_excel(writer, sheet_name=sheet_name, index=False)


def apply_standard_worksheet_style(workbook) -> None:
    if not HAS_OPENPYXL:
        raise RuntimeError("openpyxl is not available")

    header_font = Font(name="微软雅黑", bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for ws in workbook.worksheets:
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        for row in ws.iter_rows():
            values = [cell.value for cell in row]
            if values and values[0] and all(value in (None, "") for value in values[1:]):
                row[0].font = header_font
                row[0].fill = header_fill
                row[0].alignment = Alignment(horizontal="left")
        for column in ws.columns:
            max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column)
            ws.column_dimensions[column[0].column_letter].width = min(55, max_length + 5)
            for cell in column:
                cell.border = thin_border
